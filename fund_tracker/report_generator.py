"""
report_generator.py
--------------------
Builds one Fund_Report_<FundCode>.xlsx per fund with:
    - Summary        (NAV, returns, CAGR, alpha, tracking error, drawdown)
    - Holdings       (Stock, Current Weight, Previous Weight, Drift, Sector)
    - Attribution    (Stock-level and Sector-level contribution)
    - Rebalancing    (drift flags vs the prior month-end snapshot)
    - Charts sheet   (NAV vs Benchmark, Growth, Drawdown, Sector Allocation,
                       Current vs Previous Month Allocation, Contribution by Stock)

Chart images are rendered once with matplotlib, saved to
Outputs/Performance_Charts/ (per the project's stated folder layout) and
then embedded into the workbook, since openpyxl's native charts are
harder to style consistently than a matplotlib PNG.
"""

from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # headless rendering, no display needed
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import Reference  # kept for optional native-chart use
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import performance

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
FLAG_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")

PCT_FMT = "0.00%"
NUM_FMT = "#,##0.0000"


# ---------------------------------------------------------------------------
# Chart rendering (matplotlib -> PNG, both saved to disk and returned as paths)
# ---------------------------------------------------------------------------
def _save_chart(fig, fund_code: str, name: str) -> Path:
    path = config.CHARTS_DIR / f"{fund_code}_{name}.png"
    fig.savefig(path, dpi=140, bbox_inches="tight")
    plt.close(fig)
    return path


def chart_nav_vs_benchmark(daily: pd.DataFrame, fund_code: str, name: str = "nav_vs_benchmark",
                            title: str = "NAV vs Benchmark") -> Path:
    """
    Portfolio NAV (per-unit, typically tens-to-hundreds) and Benchmark NAV
    (an index level, often thousands-to-tens-of-thousands - e.g. a TRI
    index) live on very different scales. Plotting both on one shared
    y-axis makes the Portfolio NAV line look like a flat line at zero
    (it isn't - it's just tiny next to the benchmark's scale). A twin
    y-axis keeps the two series' actual shapes visible on their own scale.
    """
    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax2 = ax.twinx()

    line1, = ax.plot(daily["Date"], daily["Portfolio NAV"], label="Portfolio NAV", color="#1F4E78")
    line2, = ax2.plot(daily["Date"], daily["Benchmark NAV"], label="Benchmark NAV",
                       color="#C0392B", linestyle="--")

    ax.set_ylabel("Portfolio NAV", color="#1F4E78")
    ax2.set_ylabel("Benchmark NAV", color="#C0392B")
    ax.tick_params(axis="y", labelcolor="#1F4E78")
    ax2.tick_params(axis="y", labelcolor="#C0392B")

    ax.set_title(title)
    ax.legend(handles=[line1, line2], loc="upper left")
    ax.grid(alpha=0.3)
    fig.autofmt_xdate()
    return _save_chart(fig, fund_code, name)


def chart_growth(daily: pd.DataFrame, fund_code: str, name: str = "growth",
                  title: str = "Portfolio Growth (Cumulative Return)") -> Path:
    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax.plot(daily["Date"], daily["Cumulative Portfolio Return"], label="Portfolio", color="#1F4E78")
    ax.plot(daily["Date"], daily["Cumulative Benchmark Return"], label="Benchmark", color="#C0392B", linestyle="--")
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1.0))
    ax.set_title(title)
    ax.legend()
    ax.grid(alpha=0.3)
    fig.autofmt_xdate()
    return _save_chart(fig, fund_code, name)


def chart_drawdown(daily: pd.DataFrame, fund_code: str) -> Path:
    fig, ax = plt.subplots(figsize=(7, 3))
    ax.fill_between(daily["Date"], daily["Drawdown"], 0, color="#C0392B", alpha=0.5)
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1.0))
    ax.set_title("Drawdown")
    ax.grid(alpha=0.3)
    fig.autofmt_xdate()
    return _save_chart(fig, fund_code, "drawdown")


def chart_period_pair(period_perf: dict, fund_code: str) -> list:
    """
    For every period in performance.PERIOD_ORDER that has data available
    (see performance.compute_multi_period_performance's "Daily" key), build
    that period's OWN NAV-vs-benchmark and growth charts - i.e. a "Last 1Y"
    chart is drawn only from the last 1 year of data (rebased to 0% at that
    window's start), not sliced out of / rescaled from the since-inception
    chart. Returns a flat list of (title, Path) pairs in PERIOD_ORDER.
    Periods with no data (e.g. "Last 5Y" for a 2-year-old fund) are skipped.
    """
    charts = []
    for label in performance.PERIOD_ORDER:
        ps = period_perf.get(label)
        if not ps or not ps.get("Available") or "Daily" not in ps:
            continue
        safe_name = label.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_")
        nav_path = chart_nav_vs_benchmark(
            ps["Daily"], fund_code, name=f"period_{safe_name}_nav", title=f"NAV vs Benchmark - {label}"
        )
        growth_path = chart_growth(
            ps["Daily"], fund_code, name=f"period_{safe_name}_growth", title=f"Growth - {label}"
        )
        charts.append((f"{label} - NAV vs Benchmark", nav_path))
        charts.append((f"{label} - Growth", growth_path))
    return charts


def chart_sector_allocation(sector_contrib: pd.DataFrame, fund_code: str) -> Path:
    fig, ax = plt.subplots(figsize=(5, 5))
    ax.pie(
        sector_contrib["Weight"],
        labels=sector_contrib["Sector"],
        autopct="%1.1f%%",
        startangle=90,
        colors=plt.cm.tab20.colors,
    )
    ax.set_title("Sector Allocation")
    return _save_chart(fig, fund_code, "sector_allocation")


def chart_current_vs_previous(rebalance_df: pd.DataFrame, fund_code: str) -> Path:
    """Bar chart of each holding's Current Weight next to its Previous
    month-end Weight (see rebalance.compute_weight_drift)."""
    fig, ax = plt.subplots(figsize=(7, 3.5))
    x = range(len(rebalance_df))
    width = 0.35
    ax.bar([i - width / 2 for i in x], rebalance_df["Current Weight"], width, label="Current", color="#1F4E78")
    ax.bar([i + width / 2 for i in x], rebalance_df["Previous Weight"], width, label="Previous Month", color="#F0A500")
    ax.set_xticks(list(x))
    ax.set_xticklabels(rebalance_df["Stock Name"], rotation=45, ha="right", fontsize=7)
    ax.set_title("Current vs Previous Month Allocation")
    ax.legend()
    ax.grid(alpha=0.3, axis="y")
    return _save_chart(fig, fund_code, "current_vs_previous")


def chart_contribution_by_stock(stock_contrib: pd.DataFrame, fund_code: str) -> Path:
    df = stock_contrib.sort_values("Contribution")
    fig, ax = plt.subplots(figsize=(7, max(3, 0.35 * len(df))))
    colors = ["#C0392B" if v < 0 else "#1F4E78" for v in df["Contribution"]]
    ax.barh(df["Stock Name"], df["Contribution"], color=colors)
    ax.xaxis.set_major_formatter(mticker.PercentFormatter(xmax=1.0))
    ax.set_title("Contribution by Stock")
    ax.grid(alpha=0.3, axis="x")
    return _save_chart(fig, fund_code, "contribution_by_stock")


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------
def _style_header(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_df(ws, df: pd.DataFrame, start_row=1, start_col=1, pct_cols=None, num_cols=None):
    pct_cols = pct_cols or []
    num_cols = num_cols or []
    for j, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col_name)
    _style_header(ws, start_row, len(df.columns))

    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, col_name in enumerate(df.columns):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=row[col_name])
            cell.font = BODY_FONT
            if col_name in pct_cols:
                cell.number_format = PCT_FMT
            elif col_name in num_cols:
                cell.number_format = NUM_FMT

    for j, col_name in enumerate(df.columns):
        col_letter = get_column_letter(start_col + j)
        width = max(12, min(32, len(str(col_name)) + 4, df[col_name].astype(str).str.len().max() + 4
                             if len(df) else 12))
        ws.column_dimensions[col_letter].width = width

    return start_row + len(df) + 1  # next free row


def _flag_rebalance_rows(ws, df: pd.DataFrame, header_row: int, flag_col_name="Rebalance Required"):
    if flag_col_name not in df.columns:
        return
    col_idx = list(df.columns).index(flag_col_name) + 1
    for i, val in enumerate(df[flag_col_name], start=1):
        if val:
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=header_row + i, column=c).fill = FLAG_FILL


def _add_image(ws, path: Path, anchor: str):
    img = XLImage(str(path))
    img.width, img.height = img.width * 0.72, img.height * 0.72
    ws.add_image(img, anchor)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def build_fund_report(
    fund_code: str,
    fund_name: str,
    perf: dict,
    holdings_enriched: pd.DataFrame,
    stock_contrib: pd.DataFrame,
    sector_contrib: pd.DataFrame,
    rebalance_df: pd.DataFrame,
    threshold: float,
    period_perf: dict = None,
) -> Path:
    wb = Workbook()

    # ---- Summary ----------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{fund_name} ({fund_code}) - Performance Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    s = perf["summary"]
    cagr_annualised = s.get("CAGR Annualised", True)
    summary_rows = [
        ("Analysis Window", f"{s['Start Date'].date()} to {s['End Date'].date()}"),
        ("Portfolio NAV (Start)", s["Portfolio NAV (Start)"]),
        ("Portfolio NAV (End)", s["Portfolio NAV (End)"]),
        ("Benchmark NAV (Start)", s["Benchmark NAV (Start)"]),
        ("Benchmark NAV (End)", s["Benchmark NAV (End)"]),
        ("Absolute Return", s["Absolute Return"]),
        # CAGR/Alpha are NaN (by design - see performance.cagr()) whenever
        # this window is under ~1 year: annualising a short window massively
        # inflates or deflates whatever the short-window return happened to
        # be (a real +10% over 3 months would show as a ~46% "CAGR"). Show
        # a clear note instead of a raw NaN in that case, rather than a
        # misleading annualised number.
        ("CAGR", s["CAGR"] if cagr_annualised else "N/A (window < 1 year - see Absolute Return)"),
        ("Benchmark Return", s["Benchmark Return"]),
        ("Active Return", s["Active Return"]),
        ("Alpha", s["Alpha"] if cagr_annualised else "N/A (window < 1 year)"),
        ("Tracking Error", s["Tracking Error"]),
        ("Information Ratio", s["Information Ratio"]),
        ("Maximum Drawdown", s["Maximum Drawdown"]),
        ("Rebalance Threshold Used", threshold / 100.0),
    ]
    pct_labels = {"Absolute Return", "CAGR", "Benchmark Return", "Active Return", "Alpha",
                  "Tracking Error", "Maximum Drawdown", "Rebalance Threshold Used"}
    row = 3
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = BODY_FONT
        if label in pct_labels and isinstance(value, (int, float)):
            cell.number_format = PCT_FMT
        elif isinstance(value, float):
            cell.number_format = NUM_FMT
        row += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # ---- Performance by period (Since Inception / 5Y / 3Y / 1Y / Current FY) --
    if period_perf:
        period_header_row = row + 2
        ws.cell(row=period_header_row, column=1, value="Performance by Period").font = TITLE_FONT

        period_cols = ["Period", "Window Start", "Window End", "Absolute Return",
                       "CAGR", "Alpha", "Maximum Drawdown"]
        table_row = period_header_row + 2
        for j, col_name in enumerate(period_cols):
            ws.cell(row=table_row, column=1 + j, value=col_name)
        _style_header(ws, table_row, len(period_cols))

        pct_period_cols = {"Absolute Return", "CAGR", "Alpha", "Maximum Drawdown"}
        for i, label in enumerate(performance.PERIOD_ORDER, start=1):
            ps = period_perf.get(label, {"Available": False, "Reason": "n/a"})
            r = table_row + i
            if not ps.get("Available"):
                ws.cell(row=r, column=1, value=label).font = BODY_FONT
                ws.cell(row=r, column=2, value=ps.get("Reason", "Not available")).font = BODY_FONT
                continue
            # CAGR/Alpha are NaN (by design) whenever this period's own
            # window is under ~1 year - this is the normal case for
            # "Current FY (Apr-Mar)" for most of the financial year, and can
            # also happen to "Since Inception" for a fund under a year old.
            # Annualising a short window would otherwise blow the number up
            # (a genuine +10% over 3 months showing as a ~46% "CAGR") - show
            # "N/A" with the Absolute Return alongside it instead, which is
            # the actually-meaningful, non-annualised figure for that period.
            cagr_annualised = ps.get("CAGR Annualised", True)
            cagr_val = ps["CAGR"] if cagr_annualised else "N/A - see Absolute Return"
            alpha_val = ps["Alpha"] if cagr_annualised else "N/A"
            values = [
                label, ps["Window Start"].date().isoformat(), ps["Window End"].date().isoformat(),
                ps["Absolute Return"], cagr_val, alpha_val, ps["Maximum Drawdown"],
            ]
            for j, (col_name, val) in enumerate(zip(period_cols, values)):
                cell = ws.cell(row=r, column=1 + j, value=val)
                cell.font = BODY_FONT
                if col_name in pct_period_cols and isinstance(val, (int, float)):
                    cell.number_format = PCT_FMT
            if ps.get("Truncated"):
                ws.cell(row=r, column=1).font = Font(name="Arial", italic=True, size=10)
            if not cagr_annualised:
                ws.cell(row=r, column=1).font = Font(name="Arial", italic=True, size=10,
                                                       color="806000")

        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = max(
                ws.column_dimensions[col_letter].width or 0, 18
            )

    # ---- Holdings -----------------------------------------------------------
    ws_h = wb.create_sheet("Holdings")
    holdings_display = holdings_enriched.merge(
        rebalance_df[["ISIN", "Previous Weight", "Drift"]], on="ISIN", how="left"
    )[["Stock Name", "ISIN", "Current Weight", "Previous Weight", "Drift", "Sector", "Industry"]]
    next_row = _write_df(ws_h, holdings_display, pct_cols=[], num_cols=[])

    # ---- Attribution --------------------------------------------------------
    ws_a = wb.create_sheet("Attribution")
    ws_a["A1"] = "Stock-Level Contribution"
    ws_a["A1"].font = Font(name="Arial", bold=True, size=12)
    stock_disp = stock_contrib.copy()
    stock_disp["Current Weight"] = stock_disp["Current Weight"] / 100.0
    next_row = _write_df(ws_a, stock_disp, start_row=3,
                          pct_cols=["Current Weight", "Stock Return", "Contribution"])

    ws_a.cell(row=next_row + 1, column=1, value="Sector-Level Contribution").font = \
        Font(name="Arial", bold=True, size=12)
    sector_disp = sector_contrib.copy()
    sector_disp["Weight"] = sector_disp["Weight"] / 100.0
    _write_df(ws_a, sector_disp, start_row=next_row + 3, pct_cols=["Weight", "Contribution"])

    # ---- Rebalancing ----------------------------------------------------------
    ws_r = wb.create_sheet("Rebalancing")
    ws_r["A1"] = f"Rebalancing - drift vs previous month-end (threshold = {threshold:.1f} percentage points)"
    ws_r["A1"].font = Font(name="Arial", bold=True, size=12)
    header_row = 3
    _write_df(ws_r, rebalance_df, start_row=header_row)
    _flag_rebalance_rows(ws_r, rebalance_df, header_row)

    # ---- Charts -----------------------------------------------------------
    ws_c = wb.create_sheet("Charts")
    daily = perf["daily"]
    chart_paths = [
        chart_nav_vs_benchmark(daily, fund_code),
        chart_growth(daily, fund_code),
        chart_drawdown(daily, fund_code),
        chart_sector_allocation(sector_contrib, fund_code),
        chart_current_vs_previous(rebalance_df, fund_code),
        chart_contribution_by_stock(stock_contrib, fund_code),
    ]
    anchors = ["A1", "A21", "A41", "K1", "K21", "K41"]
    for path, anchor in zip(chart_paths, anchors):
        _add_image(ws_c, path, anchor)

    # ---- Period Charts (5Y / 3Y / 1Y / Current FY, each on its own window) --
    # Each of these is computed fresh over just that period's window (see
    # performance.compute_multi_period_performance's "Daily" key) rather than
    # sliced/rescaled out of the since-inception charts above, so a period's
    # cumulative return correctly starts at 0% at that period's own start date.
    if period_perf:
        period_charts = chart_period_pair(period_perf, fund_code)
        if period_charts:
            ws_p = wb.create_sheet("Period Charts")
            ws_p["A1"] = "Performance by Period - NAV vs Benchmark and Growth, each over its own window"
            ws_p["A1"].font = TITLE_FONT
            row_anchor = 3
            for i in range(0, len(period_charts), 2):
                pair = period_charts[i:i + 2]
                for j, (title, path) in enumerate(pair):
                    col = "A" if j == 0 else "K"
                    ws_p.cell(row=row_anchor, column=1 if j == 0 else 11, value=title).font = \
                        Font(name="Arial", bold=True, size=10)
                    _add_image(ws_p, path, f"{col}{row_anchor + 1}")
                row_anchor += 20

    # ---- save ---------------------------------------------------------------
    out_path = config.OUTPUT_DIR / f"Fund_Report_{fund_code}.xlsx"
    wb.save(out_path)
    return out_path
